#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import csv
import random
import itertools
import numpy as np

from rdkit import Chem
from rdkit import DataStructs
from rdkit.Chem import MACCSkeys
from rdkit.Chem import rdMolDescriptors
from rdkit.Chem.MACCSkeys import GenMACCSKeys

import joblib
from sklearn.model_selection import KFold
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import StratifiedKFold
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.utils.class_weight import compute_class_weight

from openpyxl import Workbook
from openpyxl import load_workbook
import statistics

import matplotlib as mpl 
mpl.use('TkAgg')
import matplotlib.pyplot as plt

from DeepNeuralNetwork import train_DNN, test_DNN, k_fold_DNN

"""
This code was written in Python 3.6.0 on a MacBook running macOS Sierra, version 10.12.
Its purpose is to evaluate different sampling techniques by using the sampled or 
unsampled data to train a Random Forest Model. 

Given a dataset a specified sampling technique is performed. 
This is followed by a hyper parameter optimization to increase the model performance.
Using the optimal hyper parameters a 10-fold cross validation is performed.
The model reaching the highest sum of performance measures is used in the external validation.
The results of both the external and internal validation are saved as .xslx tables.
If at the end of execution of this code the performance table of every sampling method
(including "nosampling") is present, a .xlsx table is created combining all the 
performance tables. Additional, plots of the internal and external performance measures
are created.

Global Variables used to choose sampling method and data set.
dataSet:    can be set as "Dili"(Drug induced liver toxicity),"AhR","LBD" or "HSE" (data sets of the tox21 challenge).
sampling:   "nosampling","augRandUS","randUS","augRandOS","randOS","kMedoids1","kMedoids2","SMOTETC" or "SMOTEVDM" can be chosen.

augRandUS:  modified random undersampling based on a fingerprint containing the most common features.
randUS:     random undersampling.
augRandOS:  modified random undersampling based on a fingerprint containing the most common features.
randOS:     random oversampling
kMedoids:   undersampling by clustering the majority data set utilitzing kMedoid algorithm.
                kMedoids1: in each step of the iteration a new set of medoids is randomly chosen.
                kMedoids2: initial set of medoids is randomly chosen, each medoid is than exchanged with 30 randomly chosen non-medoids.
SMOTE:      Implementation of the SMOTE-N introduced in [N. V. Chawla, K. W. Bowyer, L. O. Hall and W. P. Kegelmeyer (2002) "SMOTE: Synthetic Minority Over-sampling Technique"]
                SMOTETC:  Tanimoto Coefficient is used as distance/similarity metric
                SMOTEVDM: modified Value Distance Metric proposed by Cost and Salzberg (1993)

"""

global dataSet
global sampling
dataSet = "AhR" 
sampling = "kMedoids2" 

#=======#
# Input #
#=======# 
def SUBSETS(dataSet):
    """
    Reads training data set and divides data points into majority and minority
    class which are used for sampling.
    """
    inact = []
    inactSmiles = []
    inactNames = []
    act = []
    actSmiles = []
    actNames = []
    i = 0
    path = os.path.dirname(os.path.realpath(__file__)) + ""
    if dataSet == "Dili":
        actInd = 2
        smilesInd = 1
        namesInd = 0
    else:
        actInd = 1
        smilesInd = 2
        namesInd = 0
    with open("{}/Original Data Sets/{}-Train-Set.csv".format(path,dataSet),"r") as file:
            reader = csv.reader(file,delimiter=",")
            for row in reader:
                if row[actInd] == "0":
                    inact.append(i)             
                    inactSmiles.append(row[smilesInd])
                    inactNames.append(row[namesInd])
                elif row[actInd] == "1":
                    act.append(i)              
                    actSmiles.append(row[smilesInd]) 
                    actNames.append(row[namesInd])   
                i+=1
    if len(inact) < len(act):
        majInds = act[:]
        majSmiles = actSmiles[:]
        majNames = actNames[:]
        minInds = inact[:]
        minSmiles = inactSmiles[:]
        minNames = inactNames[:]
        majClass = "1"
        minClass = "0"
    else:
        majInds = inact[:]
        majSmiles = inactSmiles[:]
        majNames = inactNames[:]
        minInds = act[:]
        minSmiles = actSmiles[:]
        minNames = actNames[:]
        majClass = "0"
        minClass = "1"
    return(majClass,minClass,majInds,minInds,majSmiles,minSmiles,majNames,minNames)
    
     
def GET_FP(smiles):
    """
    Creates MACCS fingerprint from SMILES using the rdkit library.
    """
    compound = Chem.MolFromSmiles(smiles)
    fp = []
    noFp = 0
    if compound != None:
        maccs = MACCSkeys.GenMACCSKeys(compound)
        fp = np.zeros((1,))
        DataStructs.ConvertToNumpyArray(maccs, fp)
        fp = [int('%d' % (x)) for x in fp]
    else:
        noFp = 1
    return(fp,noFp)


def GET_FPS(names,inds,smiles):
    """
    Produces MACCS fingerprints for the complete data set.
    """
    fps = []
    probSmilesInd = []
    probSmPr = 0
    for ind,smi in enumerate(smiles):
        fp,noFp = GET_FP(smi)
        if noFp == 1:
            if probSmPr == 0:
                print("\nProblematic SMILES")
                probSmPr = 1
            print("Name: {}, Index in Original Dataset: {}, Smiles: {}".format(names[ind],inds[ind],smiles[ind]))
            probSmilesInd.append(ind)
        else:
            fps.append(fp)  
    for i,x in enumerate(probSmilesInd):
        del names[x-i]
        del inds[x-i]
    return(names,inds,fps)


def READ_TRAIN(dir1,file1,file2,dir2=0):
    """
    Reads sampled and unsampled data.
    """
    trainFps = []
    trainAct = []
    with open(dir1 + file1,"r") as file:
        reader = csv.reader(file,delimiter=";")
        for row in reader:
            trainFps.append([int(x) for x in row[1].split(",")])
            trainAct.append(int(row[2]))
    if dir2 == 0:
        with open(dir1 + file2,"r") as file:
            reader = csv.reader(file,delimiter=";")
            for row in reader:
                trainFps.append([int(x) for x in row[1].split(",")])
                trainAct.append(int(row[2]))
    else:
        with open(dir2 + file2,"r") as file:
            reader = csv.reader(file,delimiter=";")
            for row in reader:
                trainFps.append([int(x) for x in row[1].split(",")])
                trainAct.append(int(row[2]))
    return(trainFps,trainAct)
   
    
def GET_DATA(dataSet,sampling="test"):
    """
    Produces the "trainFps" and "trainAct" or "testFps" and "testAct" lists 
    containing fingerprints and activity of training or test data.
    """
    path = os.path.dirname(os.path.realpath(__file__))
    dirSampled = path + "/Sampled Data Sets"
    dirMinMaj = path + "/Original Data Sets"
    if sampling == "nosampling":
        fps,act = READ_TRAIN(dirMinMaj,"/{}-minClass-train.csv".format(dataSet), "/{}-majClass-train.csv".format(dataSet))
    elif sampling in ["augRandUS","randUS","kMedoids1","kMedoids2"]:
        fileName = "/{}-{}-train.csv".format(dataSet,sampling) 
        fps,act = READ_TRAIN(dirSampled,fileName,"/{}-minClass-train.csv".format(dataSet),dir2=dirMinMaj)
    elif sampling in ["SMOTETC","SMOTEVDM","augRandOS","randOS"]:
        fileName = "/{}-{}-train.csv".format(dataSet,sampling) 
        fps,act = READ_TRAIN(dirSampled,fileName,"/{}-majClass-train.csv".format(dataSet),dir2=dirMinMaj)
    elif sampling == "test":
        fps = []
        act = []
        if dataSet == "Dili":
            smilesInd = 1
            actInd = 2
        else:
            smilesInd = 2
            actInd = 1
        with open("{}/{}-Test-Set.csv".format(dirMinMaj,dataSet),"r") as file:
            reader = csv.reader(file,delimiter=",")
            if dataSet != "Dili":
                next(reader)
            for row in reader:
                fp,noFp = GET_FP(row[smilesInd])
                if noFp == 1:
                    print("nofp")
                fps.append(fp)
                act.append(int(row[actInd]))
    return(fps,act)
#================#
# Help Functions #
#================#
def MCFFP_TC(fps):
    """
    Calculates the most common feature (mcf) fingerprint for a class as well 
    as a list of similarities (Tanimoto Coefficent) between the mcf and all 
    fingerprints of the respective class.
    """
    relFreqs = []
    freqs = [0 for i in fps[0]]
    for fp in fps:
        relFreqs.append(sum(fp)/len(fp))
        for ind,val in enumerate(fp):
            freqs[ind] += val/len(fps)
    avgFreq = sum(relFreqs)/len(relFreqs)
    avgNumFeat = round(avgFreq*len(fps[0]))
    hfFeat = [(i,x) for i,x in enumerate(freqs) if x > avgFreq]
    hfFeat.sort(key=lambda x: x[1],reverse = True)
    mcfFp = [0 for i in fps[0]]
    if len(hfFeat) < avgNumFeat:
        numFeat = len(hfFeat)
    else:
        numFeat = avgNumFeat
    print(len(hfFeat),avgNumFeat)
    for i in range(0,numFeat):
        mcfFp[hfFeat[i][0]] = 1
    TCList = []
    for i in fps:
        TCList.append(TC_CALC(i,mcfFp))
    return(mcfFp,TCList)


def CHOOSE_MEDOIDS(k,oldNonMedInd,oldNonMed,oldMedInd = [],oldMed = []):
    """
    Creates a list of medoids which are used as cluster centers in both
    kMedoids methods.
    """
    meds = []
    medsInd = []
    nonMeds = []
    nonMedsInd = []
    if k > len(oldNonMedInd):
        oldNonMed = oldNonMed + oldMed
        oldNonMedInd = oldNonMedInd + oldMedInd 
        indexes = [i for i in range(0,len(oldNonMedInd))]
        samples = random.sample(indexes,k)
        bigK = 1
    else:
        indexes = [i for i in range(0,len(oldNonMedInd))]
        samples = random.sample(indexes,k)         
        bigK = 0
    for ind in indexes:
        if ind in samples:
            meds.append(oldNonMed[ind])
            medsInd.append(oldNonMedInd[ind])
        else:    
            nonMeds.append(oldNonMed[ind])
            nonMedsInd.append(oldNonMedInd[ind]) 
    if bigK == 0:
        nonMeds = nonMeds + oldMed
        nonMedsInd = nonMedsInd + oldMedInd
    return(meds,medsInd,nonMeds,nonMedsInd)


def TC_CALC(med,nonMed):
    """
    Calculates similarity (Tanimoto Coefficient) between two feature vectors.
    """
    zipList = zip(med,nonMed)
    c = len([x for x,y in zipList if x == 1 and y == 1])
    a = sum(med)
    b = sum(nonMed)
    TC = c / (a+b-c)
    return(TC)


def TC(meds,nonMeds,cluster = 0):
    """
    Calculates one or multiple similarities (Tanimoto Coefficient) between 
    medoids and non-medoids.
    If cluster is set to 1 and a list of medoids as well as a list of 
    non-medoids is given, the fingerprints are clustered by similarity.
    """
    if cluster == 0:
        if len(meds) == 167:
            medoidSims = []               
            for i in nonMeds:
                medoidSims.append(TC_CALC(meds,i))  
        else:
            medoidSims = []                   
            for i in meds:
                sims =[]                               
                for j in nonMeds:
                    sims.append(TC_CALC(i,j))           
                medoidSims.append(sims)
        return(medoidSims)
    else:
        highestSims = []
        cluster = [[] for x in range(0,k)]
        for nmFp in nonMeds:
            bestSim = 0
            for ind,mFp in enumerate(meds):
                TC = TC_CALC(mFp,nmFp)                  
                if TC > bestSim:
                    bestSim = TC
                    clusterInd = ind
            cluster[clusterInd].append(nmFp)
            highestSims.append(bestSim)
        return(cluster,highestSims)
        
    
def PROD_CLUSTER(nonMedsInd,sims,k):
    """
    Creates clusters based on similarities (Tanimoto Coefficient).
    """
    highestSims = []
    cluster = [[] for x in range(0,k)]
    for i in range(0,len(nonMedsInd)):
        bestSim = sims[0][i]
        index = 0
        for j in range(1,k):
            if sims[j][i] > bestSim:
                bestSim = sims[j][i]
                index = j
        cluster[index].append(nonMedsInd[i])
        highestSims.append(bestSim)
    return(cluster,highestSims)
     
    
def VDM_MATRIX(minFps,majFps):
    """
    Produces matrix of distances between two corresponding feature values. 
    Distances are calculated using the modified version of the 
    Value Difference Metric(VDM)[Cost and Salzberg (1993)].
    """
    lenFp = len(minFps[0])
    VDMat = [0 for i in range(0,lenFp)] 
    for featInd in range(0,lenFp):
        c1 = 0
        c2 = 0
        c1i = [0,0]
        c2i = [0,0]
        for minFp in minFps:
            if minFp[featInd] == 0:
                c1 += 1
                c1i[0] += 1
            elif minFp[featInd] == 1:
                c2 += 1
                c2i[0] += 1
        for majFp in majFps:
            if majFp[featInd] == 0:
                c1 += 1
                c1i[1] += 1
            elif majFp[featInd] == 1:
                c2 += 1
                c2i[1] += 1
        if c1 == 0 or c2 == 0:
            VDMat[featInd] = 0
        else:          
            VDMat[featInd] = abs((c1i[0]/c1 - c2i[0]/c2) + (c1i[1]/c1 - c2i[1]/c2))
    return(VDMat)
    

def VDM_CALC(minFp1,minFp2,VDMat):
    """
    Calculates distance (Value Distance Metric) between two feature vectors.
    """
    VDM = 0
    for ind,feat in enumerate(minFp1):
        if feat != minFp2[ind]:
            VDM += VDMat[ind]
    return(VDM)
    
    
def NN_CALC(fpInd,fpI,minFps,majFps,k):
    """
    Calculates the nearest neighbours by utilizing the VDM or TC.
    """
    if fpInd < k:
        iInK = 1
    else:
        iInK = 0
    if not majFps:
        # Calculation of similarities between "fpI" and every other minority class fingerprint using TC.
        nnSims = []
        inds = []
        for i in range(0,len(minFps)):
            if fpInd == i:
                continue
            TC = TC_CALC(fpI,minFps[i])   
            # First k fingerprints are added to list of nearest neighbours.
            if i < k+iInK:                                          
                nnSims.append(TC)
                inds.append(i)
                continue 
            # Nearest neighbours with minimal similarity are exchanged until all distances were calculated.    
            minim = min(nnSims)                   
            if TC > minim:                                        
                minInd = nnSims.index(minim)                      
                inds[minInd] = i   
                nnSims[minInd] = TC
    else:
        # Calculation of feature value distances.
        VDMat = VDM_MATRIX(minFps,majFps)
        # Calculation of distances between "fpI" and every other minority class fingerprint using VDM.
        nnSims = []
        inds = []
        for i in range(0,len(minFps)):
            if fpInd == i:
                continue
            VDM = VDM_CALC(fpI,minFps[i],VDMat)   
            # First k fingerprints are added to list of nearest neighbours.
            if i < k+iInK:                             
                nnSims.append(VDM)
                inds.append(i)
                continue
            # Nearest neighbours with maximal distances are exchanged until all distances were calculated. 
            maxim = max(nnSims)                                   
            if VDM < maxim:                                        
                maxInd = nnSims.index(maxim)                      
                inds[maxInd] = i
                nnSims[maxInd] = VDM
    nns = [minFps[ind] for ind in inds] 
    return(nns)                     
    

def PROD_SYN(minFp,nnsList,numUsedNNs):
    """
    Produces synthetic sample by using feature values that are most frequent 
    in "minFp" and two of its randomly chosen nearest neighbours.
    """
    usedNNs = random.sample(nnsList,numUsedNNs)
    synMin = []
    for i,feat in enumerate(minFp):
        featSum = feat
        for j in range(0,len(usedNNs)):
            featSum += usedNNs[j][i]
        featSum = featSum/(numUsedNNs+1)
        if featSum >= 0.5:
            synMin.append(1)
        else:
            synMin.append(0)
    return(synMin)


def CHECK_FOR(dataSet,sampling,bp=0,pkl=0,tab=0,all=0):
    """
    Checks if specified file exists. Also used to create filenames.
    """
    path = os.path.dirname(os.path.realpath(__file__))
    if all == 1:
        for samp in sampling:
            exist,fileName = CHECK_FOR(dataSet,samp,tab=1)
            if not exist:
                break
        return(exist)
    else:
        if bp == 1:
            directory = "{}/Best Parameters".format(path)
            fileName = "{}/{}-bestParams-{}.csv".format(directory,dataSet,sampling)
            exist = os.path.isfile(fileName)
        elif pkl == 1:
            directory = "{}/Best Rfs".format(path)
            fileName = "{}/{}-bestRf-{}.pkl".format(directory,dataSet,sampling)
            exist = os.path.isfile(fileName)
        elif tab == 1:
            directory = "{}/Validation Tables".format(path)
            fileName = "{}/{}-RF-{}-Validation-Table.xlsx".format(directory,dataSet,sampling)
            exist = os.path.isfile(fileName)        
        else:
            if sampling == "nosampling":
                directory = "{}/Original Data Sets".format(path)        
                fileName = []
                fileName.append("{}/{}-majClass-train.csv".format(directory,dataSet))
                fileName.append("{}/{}-minClass-train.csv".format(directory,dataSet))
                exist = os.path.isfile(fileName[0])
            else:
                directory = "{}/Sampled Data Sets".format(path)        
                fileName = "{}/{}-{}-train.csv".format(directory,dataSet,sampling)
                exist = os.path.isfile(fileName)
        if exist == False:
            if not os.path.exists(directory):
                os.makedirs(directory)
        return(exist,fileName)   
#==========#
# Sampling #
#==========#
def AUG_RAND_US(minInds,majFps,majNames):
    """
    Performs random undersampling that utilizes the mcf fingerpint.
    """
    newMajFps = list(majFps)
    newMajNames = list(majNames)
    while len(newMajFps) != len(minInds):
        mcfFp,TCList = MCFFP_TC(newMajFps)
        orTCList = list(TCList)
        minTcInds = []
        for i in range(0,round(len(minInds)/2)):
            minOfTc = min(TCList)
            minTcInds.append(orTCList.index(minOfTc))
            TCList.remove(minOfTc)
        delCounter = round(len(minInds)/4)    
        while len(newMajFps) != len(minInds):
            if delCounter > 0:
                rndInd = random.choice(minTcInds)
                del newMajFps[rndInd]
                del newMajNames[rndInd]
                minTcInds.remove(rndInd)
                minTcInds = [ind-1 if ind > rndInd else ind for ind in minTcInds]
                delCounter -= 1
            else:
                break
    return(newMajNames,newMajFps)
        

def AUG_RAND_OS(majInds,minFps,minNames):
    """
    Performs random oversampling that utilizes the mcf fingerpint.
    """
    newMinFps = list(minFps)
    newMinNames = list(minNames)
    while len(newMinFps) != len(majInds):
        mcfFp,TCList = MCFFP_TC(newMinFps)
        orTCList = list(TCList)
        maxTcInds = []
        for i in range(0,round(len(minInds)/2)):
            maxOfTc = max(TCList)
            maxTcInds.append(orTCList.index(maxOfTc))
            TCList.remove(maxOfTc) 
        appCounter = round(len(minInds)/4) 
        while len(newMinFps) != len(majInds):
            if appCounter > 0:
                rndInd = random.choice(maxTcInds)
                newMinFps.append(newMinFps[rndInd])
                newMinNames.append(newMinNames[rndInd])
                maxTcInds.remove(rndInd)
                appCounter -= 1
            else:
                break
    return(newMinNames,newMinFps)
    

def RAND_US(minInds,majFps,majNames):
    """
    Performs random undersampling.
    """
    newMajFps = random.sample(majFps,len(minInds))
    newMajNames = [majNames[majFps.index(fp)] for fp in newMajFps]
    return(newMajNames,newMajFps)
    
    
def RAND_OS(majInds,minFps,minNames):
    """
    Performs random oversampling.
    """    
    newMinFps = list(minFps)
    newMinNames = list(minNames)
    while len(newMinFps) != len(majInds):
        ind = random.randint(0,len(minFps)-1)
        newMinFps.append(minFps[ind])
        newMinNames.append(minNames[ind])
    return(newMinNames,newMinFps)


def KMEDOIDS_1(k,majInds,majFps):
    """
    Undersamples the majority class by clustering the datapoints. 
    The cluster centers or medoids are the undersampled majority class.
    The iteration consists of choosing random fingerprints as new 
    medoids and calculating new clusters and cost.
    The set of medoids resulting in the highest sum of similarities
    is used as undersampled data. 
    """
    # choosing first medoids, calculation of similarities,clusters and cost
    meds,medsInd,nonMeds,nonMedsInd = CHOOSE_MEDOIDS(k,majInds,majFps)    
    bestMeds = list(meds)
    bestMedsInd = list(medsInd)
    cluster,highestSims = TC(meds,nonMeds,cluster = 1)
    simSum = sum(highestSims)
    for i in range(0,100):
        # choosing new medoids, calculation of similarities,clusters and cost
        meds,medsInd,nonMeds,nonMedsInd = CHOOSE_MEDOIDS(k,nonMedsInd,nonMeds,oldMedInd = medsInd,oldMed = meds)
        cluster,highestSims = TC(meds,nonMeds,cluster = 1)
        newSimSum = sum(highestSims)
        # saving cost and medoids if cost was improved
        if newSimSum > simSum:
            simSum = newSimSum
            bestMeds = list(meds)
            bestMedsInd = list(medsInd)
    return(bestMedsInd,bestMeds)


def KMEDOIDS_2(k,majInds,majFps):
    """
    Undersamples the majority class by clustering the datapoints. 
    The cluster centers or medoids are the undersampled majority class.
    The first complete set of medoids is chosen randomly.
    The iteration consists of exchanging each medoid with a random
    non-medoid and calculating the resulting clusters and cost.
    The set of medoids resulting in the highest sum of similarities
    is used as undersampled data.
    """
    # choosing first medoids, calculation of similarities,clusters and cost
    meds,medsInd,nonMeds,nonMedsInd = CHOOSE_MEDOIDS(k,majInds,majFps)
    sims = TC(meds,nonMeds)
    cluster,highestSims = PROD_CLUSTER(nonMedsInd,sims,k) 
    simSum = sum(highestSims)
    for i in range(0,len(meds)):
        for x in range(0,30):            
            j = random.randrange(0,len(nonMedsInd))
            # saving medoid and non-medoid before exchanging them
            newMedInd = nonMedsInd[j]       
            newMed = nonMeds[j]             
            oldMedInd = medsInd[i]
            oldMed = meds[i]
            # exchanging medoid and non-medoid
            medsInd[i] = newMedInd          
            meds[i] = newMed                
            nonMedsInd[j] = oldMedInd       
            nonMeds[j] = oldMed             
            # adjusting similarities 
            newSims = list(sims)
            newMedSims = TC(newMed,nonMeds)             
            newSims[i] = newMedSims                           
            for y in range(0,k):                            
                if y == i:                                  
                    continue
                sim = TC_CALC(meds[y],oldMed)                    
                newSims[y][j] = sim
            # clustering and cost calculation based on new similarities
            newCluster,newHighestSims = PROD_CLUSTER(nonMedsInd,newSims,k)
            newSimSum = sum(newHighestSims)
            # saving new similarities, cost and clusters if previous cost was improved 
            if newSimSum > simSum:
                sims = newSims
                simSum = newSimSum
                cluster = newCluster
            # reverse exchange of medoid and non-medoid if cost was not improved
            else:
                medsInd[i] = oldMedInd
                meds[i] = oldMed
                nonMedsInd[j] = newMedInd
                nonMeds[j] = newMed
    return(medsInd,meds)
    

def SMOTE(minFps,overSam,k,numUsedNNs,majFps=[]):
    """
    Produces synthetic samples by utilizing nearest neighbors and majority 
    voting. Based on SMOTE-N introduced in 
    [N. V. Chawla, K. W. Bowyer, L. O.Hall, W. P. Kegelmeyer, 
    “SMOTE: synthetic minority over-sampling technique,” 
    Journal of artificial intelligence research, 321-357, 2002.]
    """
    nnsList = []
    # Produces neighbourhood for every minority class fingerprint.
    for i in range(0,len(minFps)):
        nns = NN_CALC(i,minFps[i],minFps,majFps,k)
        nnsList.append(nns)
        print("{} of {} neighbourhoods created".format(i+1,len(minFps)))    
    # 
    c = 0
    synMins = []
    if overSam > len(minFps):                               
        overSamC = overSam
        while overSamC > len(minFps):                       
            overSamC = overSamC - len(minFps)
            c += 1
        while overSamC != 0:                                  
            minInd = random.randrange(0,len(minFps))          
            synMins.append(PROD_SYN(minFps[minInd],nnsList[minInd],numUsedNNs))
            overSamC -= 1
    # 
    minInd = 0
    while len(synMins) != overSam:                           
        synMins.append(PROD_SYN(minFps[minInd],nnsList[minInd],numUsedNNs)) 
        minInd += 1
        if minInd == len(minFps) and c != 0:                
            minInd = 0
            c -= 1
    return(synMins) 
#========#
# Output #
#========#
def PROD_CSV(fileName,fps,names,classBit):
    """
    Produces .csv files of sampled or unsampled data. 
    """    
    with open (fileName,"w") as file:
        # Writing of sampled data.
        for ind,fp in enumerate(fps):
            fpString = ""
            for i in range(0,len(fp)):
                if i == len(fp)-1:
                    fpString = fpString + str(fp[i])
                else:
                    fpString = fpString + str(fp[i]) + ","
            outString = names[ind] + ";" + fpString +  ";" + classBit + "\n"
            file.write(outString)
#==============#
#Random Forest #
#==============#
def FIND_BEST_RF(fileName,fps,act,sampling): 
    """
    Performs an exhaustive Grid Search to find best performing 
    hyper-parameters.
    """
    fpsIndList = [i for i in range(0,len(fps))] 
    trainFpsInds, testFpsInds, trainAct, testAct = train_test_split(fpsIndList, act, test_size=0.1, stratify=act)
    trainFps = [fps[x] for i,x in enumerate(trainFpsInds)]
    testFps = [fps[x] for i,x in enumerate(testFpsInds)]
    clfRF=RandomForestClassifier()
    maxD = [i*100 for i in range(2,9)]
    nEst = 1000
    classWeights = compute_class_weight(class_weight = "balanced", classes = [0,1], y = trainAct)
    classWDict = dict(zip([0,1], trainAct))
    param_grid_RF = { 
        "max_depth": maxD,
	    "n_estimators": [nEst],
	    "max_features": ["sqrt"],
	    "class_weight": [classWDict],
	    "warm_start": [True],
        "criterion": ["gini","entropy"]} 
    rfGridSearch = GridSearchCV(estimator=clfRF, param_grid=param_grid_RF, cv=2, error_score = 'raise').fit(trainFps, trainAct)
    probas = rfGridSearch.predict_proba(testFps)
    posProbas = [probas[j][1] for j, x in enumerate(probas)]
    rocAuc = roc_auc_score(testAct,posProbas)
    bestParams = rfGridSearch.best_params_
    # Saving hyper-parameters as .csv file.    
    with open(fileName,"w") as file:
        outString = str(bestParams["n_estimators"]) + "," + str(bestParams["max_depth"]) + "," + str(bestParams["criterion"])+"," + "best ROC-AUC: " + str(rocAuc)
        file.write(outString)    
    
    
def K_FOLD(k,fps,act,dataSet,sampling):
    """
    A 10 fold cross-validation of the Random Forest.
    Returns performance measures(ROC-AUC,balanced accuracy,accuracy,
    sensitivity,specificity and F-measure),as well as the standard 
    deviations of these measures for the cross-validation. 
    The best performing model and the highest performance measure 
    values that were reached are returned as well.
    """
    skfold = StratifiedKFold(n_splits=k)
    rocAuc = []
    balAcc = []
    acc = []
    sens = []
    spec = []
    fMeasure = []
    allMeasures = []
    mlAlgList = []
    # Reading hyper-parameters.
    path = os.path.dirname(os.path.realpath(__file__))
    with open("{}/Best Parameters/{}-bestParams-{}.csv".format(path,dataSet,sampling),"r") as file:
        reader = csv.reader(file)
        firstRow = next(reader)     
        bestNe = int(firstRow[0])
        bestMd = int(firstRow[1])
        bestCrit = firstRow[2]
    # Performing k-Fold cross-validation.
    for train, test in skfold.split(fps,act):
        fps_train = [fps[x] for i,x in enumerate(train)]
        fps_test = [fps[x] for i,x in enumerate(test)]
        act_train = [act[x] for i,x in enumerate(train)]
        act_test = [act[x] for i,x in enumerate(test)]
        mlAlg = RandomForestClassifier(n_estimators=bestNe,warm_start=True,criterion=bestCrit, max_depth=bestMd)
        mlAlg.fit(fps_train,act_train)
        mlAlgList.append(mlAlg)
        result = mlAlg.predict(fps_test)
        probas = mlAlg.predict_proba(fps_test)
        # Calculation of performance measures.
        posProbas = [x[1] for x in probas]
        rightPreds = [i for i, x in enumerate(result) if x == act_test[i]]
        falsePreds = [i for i, x in enumerate(result) if x != act_test[i]]
        truePos = [x for i, x in enumerate(rightPreds) if act_test[x] == 1]
        trueNeg = [x for i, x in enumerate(rightPreds) if act_test[x] == 0]
        falsePos = [x for i, x in enumerate(falsePreds) if result[x] == 1]
        falseNeg = [x for i, x in enumerate(falsePreds) if result[x] == 0]
        active = [i for i, x in enumerate(act_test) if x==1]
        inactive = [i for i, x in enumerate(act_test) if x==0]
        rocAuc.append(roc_auc_score(act_test,posProbas))
        rec = len(truePos)/len(active)
        sens.append(rec)
        spec.append((len(trueNeg)/len(inactive)))
        balAcc.append((((len(truePos) / len(active)) + (len(trueNeg)/len(inactive)))/2))
        acc.append((len(truePos)+len(trueNeg))/len(act_test))
        if len(truePos) == 0:#+ len(falsePos))  
            fMeasure.append(0)
        else:
            prec = len(truePos)/(len(truePos)+len(falsePos))
            fMeasure.append(2*(prec*rec)/(prec+rec))
    allMeasures = [np.sum(rocAuc)/10,np.sum(balAcc)/10,np.sum(acc)/10,np.sum(sens)/10,np.sum(spec)/10,np.sum(fMeasure)/10]
    allStds = [np.std(rocAuc),np.std(balAcc),np.std(acc),np.std(sens),np.std(spec),np.std(fMeasure)]
    # Selection of best performing Random Forest.
    bestSum = 0
    for i in range(0,len(mlAlgList)):
        newSum = float(rocAuc[i]) + float(acc[i]) + float(sens[i]) + float(spec[i] + float(fMeasure[i]))
        if newSum > bestSum:
            bestSum = newSum
            bestMlAlg = mlAlgList[i]
    bestM = [max(rocAuc),max(balAcc),max(acc),max(sens),max(spec),max(fMeasure)]
    return(allMeasures,allStds,bestM,bestMlAlg)


def EXT_VAL(bestRf,testFps,testAct):
    """
    Uses best performing Random Forest for external prediction and calculates
    performance measures.
    """
    result = bestRf.predict(testFps)
    probas = bestRf.predict_proba(testFps)
    # Calculation of performance measures.
    posProbas = [x[1] for x in probas]
    rightPreds = [i for i, x in enumerate(result) if x == testAct[i]]
    falsePreds = [i for i, x in enumerate(result) if x != testAct[i]]
    truePos = [x for i, x in enumerate(rightPreds) if testAct[x] == 1]
    trueNeg = [x for i, x in enumerate(rightPreds) if testAct[x] == 0]
    falsePos = [x for i, x in enumerate(falsePreds) if result[x] == 1]
    falseNeg = [x for i, x in enumerate(falsePreds) if result[x] == 0]
    active = [i for i, x in enumerate(testAct) if x==1]
    inactive = [i for i, x in enumerate(testAct) if x==0]
    rocAuc= roc_auc_score(testAct,posProbas)
    sens = len(truePos)/len(active)
    spec = len(trueNeg)/len(inactive)
    acc = (len(truePos)+len(trueNeg))/len(testAct)
    if (len(truePos) + len(falsePos))  == 0 or len(truePos) == 0:
        fMeasure = 0
    else:
        prec = len(truePos)/(len(truePos)+len(falsePos))
        fMeasure = 2*(prec*sens)/(prec+sens)
    extMeasures = [rocAuc,acc,sens,spec,fMeasure]
    return(extMeasures)
#============#
# Validation #
#============#
def PROD_TAB(allM,allStds,bestM,extM,sampling,fileName):
    """
    Creates .xlsx file of performance measures.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sampling
    ws["B1"] = "int. ROC-AUC"
    ws["C1"] = "ex. ROC-AUC"
    ws["D1"] = "Bal-Acc"
    ws["E1"] = "int. Acc"
    ws["F1"] = "ex. Acc"
    ws["G1"] = "int. Sens"
    ws["H1"] = "ex. Sens"
    ws["I1"] = "int. Spec"
    ws["J1"] = "ex. Spec"
    ws["K1"] = "int. F-Measure"
    ws["L1"] = "ex. F-Measure"
    ws["B{}".format(2)] = "%.3f" % allM[0] + "+/-" + "%.3f" % allStds[0] + " | " + "best:" + "%.3f" % bestM[0]
    ws["C{}".format(2)] = "%.3f" % extM[0]
    ws["D{}".format(2)] = "%.3f" % allM[1] + "+/-" + "%.3f" % allStds[1] + " | " + "best:" + "%.3f" % bestM[1]
    ws["E{}".format(2)] = "%.3f" % allM[2] + "+/-" + "%.3f" % allStds[2] + " | " + "best:" + "%.3f" % bestM[2]
    ws["F{}".format(2)] = "%.3f" % extM[1]
    ws["G{}".format(2)] = "%.3f" % allM[3] + "+/-" + "%.3f" % allStds[3] + " | " + "best:" + "%.3f" % bestM[3]
    ws["H{}".format(2)] = "%.3f" % extM[2]
    ws["I{}".format(2)] = "%.3f" % allM[4] + "+/-" + "%.3f" % allStds[4] + " | " + "best:" + "%.3f" % bestM[4]
    ws["J{}".format(2)] = "%.3f" % extM[3]
    ws["K{}".format(2)] = "%.3f" % allM[5] + "+/-" + "%.3f" % allStds[5] + " | " + "best:" + "%.3f" % bestM[5]
    ws["L{}".format(2)] = "%.3f" % extM[4]
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 25
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 12
    wb.save(fileName)
    

def READ_TAB(fileName,comTab=0,samName=0):
    """
    Reads existing tables of performance measures.
    """
    wb = load_workbook(fileName)
    cols = ["B2","C2","D2","E2","F2","G2","H2","I2","J2","K2","L2"]
    if comTab == 0:
        ws = wb[sampling]
        allM = []
        allStds = []
        bestM = []
        extMs = []
        for col in cols:
            val = ws[col].value
            if col in ["B2","D2","E2","G2","I2","K2"]:
                valSplit1 = val.split("+/-")
                valSplit2 = valSplit1[1].split("|")
                allM = valSplit1[0]
                allStds = valSplit2[0]
                bestM = valSplit2[1]
            else:
                extMs.append(val)
        return(allM,allStds,bestM,extMs)
    elif comTab == 1:
        ws = wb[samName]
        perfMs = []
        for col in cols:
            perfMs.append(ws[col].value)
        return(perfMs)
    
    
def GET_ALL_DATA(dataSet,samNames = []):
    """
    Collects performance measures of all sampling methods.
    """
    path = os.path.dirname(os.path.realpath(__file__))
    if not samNames:
        samNames = ["nosampling","augRandUS","randUS","augRandOS","randOS","kMedoids1","kMedoids2","SMOTETC","SMOTEVDM"]
        allMs = []
        for samName in samNames:
            fileName = "{}/Validation Tables/{}-RF-{}-Validation-Table.xlsx".format(path,dataSet,samName)
            perfMs = READ_TAB(fileName,comTab=1,samName=samName)
            allMs.append(perfMs)
    else:
        fileName = "{}/Validation Tables/{}-RF-{}-Validation-Table.xlsx".format(path,dataSet,samNames)
        perfMs = READ_TAB(fileName,comTab=1,samName=samNames)
        allMs = perfMs    
    return(allMs)     


def PROD_COM_TAB(dataSet):
    """
    Produces table containing performance measures of all sampling methods.
    """
    allMs = GET_ALL_DATA(dataSet)
    samNames = ["nosampling","augRandUS","randUS","augRandOS","randOS","kMedoids1","kMedoids2","SMOTETC","SMOTEVDM"]
    wb = Workbook()
    ws = wb.active
    ws["B1"] = "int. ROC-AUC"
    ws["C1"] = "ex. ROC-AUC"
    ws["D1"] = "Bal-Acc"
    ws["E1"] = "int. Acc"
    ws["F1"] = "ex. Acc"
    ws["G1"] = "int. Sens"
    ws["H1"] = "ex. Sens"
    ws["I1"] = "int. Spec"
    ws["J1"] = "ex. Spec"
    ws["K1"] = "int. F-Measure"
    ws["L1"] = "ex. F-Measure"
    for i in range(0,len(samNames)):
        ws["A{}".format(i+2)] = samNames[i]
        ws["B{}".format(i+2)] = allMs[i][0]
        ws["C{}".format(i+2)] = allMs[i][1]
        ws["D{}".format(i+2)] = allMs[i][2]
        ws["E{}".format(i+2)] = allMs[i][3]
        ws["F{}".format(i+2)] = allMs[i][4]
        ws["G{}".format(i+2)] = allMs[i][5]
        ws["H{}".format(i+2)] = allMs[i][6]
        ws["I{}".format(i+2)] = allMs[i][7]
        ws["J{}".format(i+2)] = allMs[i][8]
        ws["K{}".format(i+2)] = allMs[i][9]
        ws["L{}".format(i+2)] = allMs[i][10]
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 25
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 12

    path = os.path.dirname(os.path.realpath(__file__))
    wb.save("{}/Validation Tables/{}-Complete-Validation-Table.xlsx".format(path,dataSet))
         

def GET_Ms(dataSet,plotName="",samName="",allM=0):
    """
    Collects performance measures to be used in plots. 
    """
    if not samName:
        allMs = GET_ALL_DATA(dataSet)
    else:
        allMs = GET_ALL_DATA(dataSet,samNames=samName)
    if allM == 0:
        # Collecting specific measure for all sampling methods.
        if plotName == "roc":
            intAvg = []
            intStd = []
            ext = []
            for ms in allMs:
                splitMs = ms[0].split("+/-")
                secSplitMs = splitMs[1].split("|")
                intAvg.append(float(splitMs[0]))
                intStd.append(float(secSplitMs[0]))
                ext.append(float(ms[1]))
            mName = "ROC-AUC"
        elif plotName == "acc":
            intAvg = []
            intStd = []
            ext = []
            for ms in allMs:
                splitMs = ms[3].split("+/-")
                secSplitMs = splitMs[1].split("|")
                intAvg.append(float(splitMs[0]))
                intStd.append(float(secSplitMs[0]))
                ext.append(float(ms[4]))
            mName = "Accuracy"
        elif plotName == "sens":
            intAvg = []
            intStd = []
            ext = []
            for ms in allMs:
                splitMs = ms[5].split("+/-")
                secSplitMs = splitMs[1].split("|")
                intAvg.append(float(splitMs[0]))
                intStd.append(float(secSplitMs[0]))
                ext.append(float(ms[6]))
            mName = "Sensitivity"
        elif plotName == "spec":
            intAvg = []
            intStd = []
            ext = []
            for ms in allMs:
                splitMs = ms[7].split("+/-")
                secSplitMs = splitMs[1].split("|")
                intAvg.append(float(splitMs[0]))
                intStd.append(float(secSplitMs[0]))
                ext.append(float(ms[8]))
            mName = "Specificity"
        elif plotName == "fm":
            intAvg = []
            intStd = []
            ext = []
            for ms in allMs:
                splitMs = ms[9].split("+/-")
                secSplitMs = splitMs[1].split("|")
                intAvg.append(float(splitMs[0]))
                intStd.append(float(secSplitMs[0]))
                ext.append(float(ms[10]))
            mName = "F-Measure"
        return(intAvg,intStd,ext,mName)
    elif allM == 1:
        # Collecting all internal measures.
        if not samName:
            allIntM = []
            allIntStd = []
            msInd = [0,3,5,7]
            for ms in allMs:
                splitMs = []
                stds = []
                for i in msInd:
                    splitM = ms[i].split("+/-")
                    secSplitMs = splitM[1].split("|")
                    splitMs.append(float(splitM[0]))
                    stds.append(float(secSplitMs[0]))
                allIntM.append(splitMs)
                allIntStd.append(stds)
            return(allIntM,allIntStd)
        else:
            splitMs = []
            stds = []
            msInd = [0,2,3,5,7,9]
            for i in msInd:
                splitM = allMs[i].split("+/-")
                splitMs.append(float(splitM[0]))
            allIntM = splitMs
        return(allIntM)
    elif allM == 2:
        # Collecting all external measures.
        allExM = []
        msInd = [1,4,6,8]
        for ms in allMs:
            exM = []
            for i in msInd:
                exM.append(float(ms[i]))
            allExM.append(exM)
        return(allExM)

    
def PROD_PLOTS(dataSet):
    """
    Produces plots of permormance measure.
    """
    plotNames = ["roc","acc","sens","spec","fm"]
    for plotName in plotNames:
        intAvg,intStd,ext,mName = GET_Ms(dataSet,plotName=plotName)
        fig, ax = plt.subplots()
        ax.set_title("{} for target {}".format(mName,dataSet),loc='left')
        index = np.arange(len(intAvg))
        bar_width = 0.35
        bar1 = ax.bar(index, intAvg, bar_width,color='0.3',yerr=intStd,label='Internal')
        bar2 = ax.bar(index + bar_width, ext, bar_width,color='grey',label='External')
        ax.set_xlabel('Sampling Methods')
        ax.set_ylabel('Score')
        ax.set_xticks(index + bar_width / 2)
        ax.set_xticklabels(("nosampling","augRandUS","randUS","augRandOS","randOS","kMedoids1","kMedoids2","SMOTETC","SMOTEVDM"))
        ax.legend(bbox_to_anchor=(1, 1.14), loc=1, borderaxespad=0.)
        ax.autoscale_view()
        fig.autofmt_xdate()
        path = os.path.dirname(os.path.realpath(__file__))
        directory = "{}/Validation Plots".format(path)
        if not os.path.exists(directory):
            os.makedirs(directory)
        fileName =  "{}/{}-{}-plot.png".format(directory,dataSet,mName)    
        plt.savefig(fileName)


def PROD_ALLM_PLOT(dataSet,allM = 1):
    """
    Produces plot of all measures for all sampling methods for one dataset.
    """    
    fig, ax = plt.subplots()
    index = np.arange(4)
    bar_width = 0.1
    if allM == 1:
        allIntM,allIntStd = GET_Ms(dataSet,allM=1)
        ax.set_title("All Measures for {}".format(dataSet),loc='left')
        bar1 = ax.bar(index,allIntM[0],bar_width,color='0.3',yerr=allIntStd[0],label='nosampling')
        bar2 = ax.bar(index + bar_width, allIntM[1], bar_width,color='maroon',yerr=allIntStd[1],label='augRandUS')
        bar3 = ax.bar(index + 2*bar_width, allIntM[2], bar_width,color='firebrick',yerr=allIntStd[2],label='randUS')
        bar4 = ax.bar(index + 3*bar_width, allIntM[3], bar_width,color= 'royalblue',yerr=allIntStd[3],label='augRandOS')
        bar5 = ax.bar(index + 4*bar_width, allIntM[4], bar_width,color='blue',yerr=allIntStd[4],label='randOS')
        bar6 = ax.bar(index + 5*bar_width, allIntM[5], bar_width,color='indianred',yerr=allIntStd[5],label='kMedoids1')
        bar7 = ax.bar(index + 6*bar_width, allIntM[6], bar_width,color='lightcoral',yerr=allIntStd[6],label='kMedoids2')
        bar8 = ax.bar(index + 7*bar_width, allIntM[7], bar_width,color='mediumblue',yerr=allIntStd[7],label='SMOTE-TC')
        bar9 = ax.bar(index + 8*bar_width, allIntM[8], bar_width,color='navy',yerr=allIntStd[8],label='SMOTE-VDM')
    elif allM == 2:
        allExM = GET_Ms(dataSet,allM=2)
        ax.set_title("All External Measures for {}".format(dataSet),loc='left')
        bar1 = ax.bar(index,allExM[0],bar_width,color='0.3',label='nosampling')
        bar2 = ax.bar(index + bar_width, allExM[1], bar_width,color='maroon',label='augRandUS')
        bar3 = ax.bar(index + 2*bar_width, allExM[2], bar_width,color='firebrick',label='randUS')
        bar4 = ax.bar(index + 3*bar_width, allExM[3], bar_width,color= 'royalblue',label='augRandOS')
        bar5 = ax.bar(index + 4*bar_width, allExM[4], bar_width,color='blue',label='randOS')
        bar6 = ax.bar(index + 5*bar_width, allExM[5], bar_width,color='indianred',label='kMedoids1')
        bar7 = ax.bar(index + 6*bar_width, allExM[6], bar_width,color='lightcoral',label='kMedoids2')
        bar8 = ax.bar(index + 7*bar_width, allExM[7], bar_width,color='mediumblue',label='SMOTE-TC')
        bar9 = ax.bar(index + 8*bar_width, allExM[8], bar_width,color='navy',label='SMOTE-VDM')
    ax.set_xlabel('Measure')
    ax.set_xticks(index + 9*bar_width / 2)
    ax.set_xticklabels(("ROC-AUC","Accuracy","Sensitivity","Specificity"))
    ax.set_yticklabels(("0.0","0.2","0.4","0.6","0.8","1.0"))
    lgd = ax.legend(bbox_to_anchor=(1.265, 1), loc=1, borderaxespad=0.)
    ax.autoscale_view()
    fig.autofmt_xdate()
    path = os.path.dirname(os.path.realpath(__file__))
    if allM == 1:
        fileName = "{}/Validation Plots/{}-allM-Plot.png".format(path,dataSet)
    elif allM == 2:
        fileName = "{}/Validation Plots/{}-allExM-Plot.png.png".format(path,dataSet)
    fig.savefig(fileName, bbox_extra_artists=(lgd,), bbox_inches='tight')
#======
# 
#======
# Oriiginal data set is divided into majority and minority class.
majClass,minClass,majInds,minInds,majSmiles,minSmiles,majNames,minNames = SUBSETS(dataSet)
majNames,majInds,majFps = GET_FPS(majNames,majInds,majSmiles)
minNames,minInds,minFps = GET_FPS(minNames,minInds,minSmiles)
# Checks if specified data set already exists.
allSampling = ["nosampling","augRandUS","randUS","kMedoids1","kMedoids2","augRandOS","randOS","SMOTETC","SMOTEVDM"]
ex_data,fileName_data = CHECK_FOR(dataSet,sampling)
if not ex_data:
    print("Sampling...", end="",flush=True)
    # Creates .csv files of majority class and minority class.
    if sampling == allSampling[0]:
        PROD_CSV(fileName_data[0],majFps,majNames,majClass)
        PROD_CSV(fileName_data[1],minFps,minNames,minClass)
    # Sampling and creation of .csv file of sampled data.
    elif sampling in allSampling[1:5]:
        if sampling == "augRandUS": 
            newMajNames,newMajFps = AUG_RAND_US(minInds,majFps,majNames)
        elif sampling == "randUS":
            newMajNames,newMajFps = RAND_US(minInds,majFps,majNames)
        elif sampling == "kMedoids1": 
            k = len(minInds)
            newMajInd,newMajFps = KMEDOIDS_1(k,majInds,majFps)
            newMajNames = [majNames[majInds.index(ind)] for ind in newMajInd]
        elif sampling == "kMedoids2": 
            k = len(minInds)       
            newMajInd,newMajFps = KMEDOIDS_2(k,majInds,majFps)
            newMajNames = [majNames[majInds.index(ind)] for ind in newMajInd]
        PROD_CSV(fileName_data,newMajFps,newMajNames,majClass)
    elif sampling in allSampling[5:]:
        if sampling == "augRandOS":
            newMinNames,newMinFps = AUG_RAND_OS(majInds,minFps,minNames)
        elif sampling == "randOS":
            newMinNames,newMinFps = RAND_OS(majInds,minFps,minNames)
        elif "SMOTE" in sampling:
            overSam = len(majInds)-len(minInds)
            if "VDM" in sampling:
                synFps = SMOTE(minFps,overSam,5,2,majFps=majFps)
            else:
                synFps = SMOTE(minFps,overSam,5,2)
            synNames = ["synFps{}".format(i) for i in range(0,len(synFps))]
            newMinFps = minFps + synFps
            newMinNames = minNames + synNames
        PROD_CSV(fileName_data,newMinFps,newMinNames,minClass)
    print("Done", end="\n")

# Get Data
trainFps, trainAct = GET_DATA(dataSet,sampling=sampling)

# Hyper-Parameter optimization.
ex_bp,fileName_bp = CHECK_FOR(dataSet,sampling,bp=1)
if not ex_bp:
    print("Hyper-Parameter optimization...", end="",flush=True)
    FIND_BEST_RF(fileName_bp,trainFps,trainAct,sampling)
    print("Done", end="\n")
# Performing k-Fold cross-validation of Random Forest.
testFps,testAct = GET_DATA(dataSet)
ex_pkl,fileName_pkl = CHECK_FOR(dataSet,sampling,pkl=1)
if not ex_pkl:
    print("Cross-Validation...", end="",flush=True)
    allM,allStds,bestM,bestRf = K_FOLD(10,trainFps,trainAct,dataSet,sampling)
    joblib.dump(bestRf, fileName_pkl) 
    print("Done", end="\n")
else:
    bestRf = joblib.load(fileName_pkl)
    allM = GET_Ms(dataSet,samName=sampling,allM=1)
# Prediction of test set with best performing Random Forest
print("External Validation...", end=" ")
extMs = EXT_VAL(bestRf,testFps,testAct)
print("Done", end="\n")
# Creation of validation table.
ex_tab,fileName_tab = CHECK_FOR(dataSet,sampling,tab=1)
if not ex_tab:
    print("Creation of validation table...", end="",flush=True)
    PROD_TAB(allM,allStds,bestM,extMs,sampling,fileName_tab)  
    print("Done", end="\n")
ex_allTab = CHECK_FOR(dataSet,allSampling,all=1)
if ex_allTab:
    print("Creation of complete validation table...", end="",flush=True)
    PROD_COM_TAB(dataSet)
    print("Done", end="\n")
    print("Creation of performance measure plots...", end="",flush=True)
    PROD_PLOTS(dataSet)
    PROD_ALLM_PLOT(dataSet,allM = 1)
    PROD_ALLM_PLOT(dataSet,allM = 2)
    print("Done", end="\n")

# Deep Neural Network model
DNN_model = train_DNN(trainFps, trainAct)
joblib.dump(DNN_model, dataSet+'-'+'DNN'+'-'+sampling+'.pkl') 
testFps, testAct = GET_DATA(dataSet, sampling=sampling)
kfold_res = k_fold_DNN(10, trainFps, trainAct)
DNN_ext_res, ext_compare = test_DNN(DNN_model, testFps, testAct)
dnn_mean = kfold_res.drop(["FNR", "FPR", "AUC_PR"], axis=1).apply(statistics.mean, axis=0).to_list()
dnn_stdev = kfold_res.drop(["FNR", "FPR", "AUC_PR"], axis=1).apply(statistics.stdev, axis=0).to_list()
dnn_max = kfold_res.drop(["FNR", "FPR", "AUC_PR"], axis=1).apply(max, axis=0).to_list()
PROD_TAB(dnn_mean, dnn_stdev, dnn_max, DNN_ext_res, sampling, dataSet+'-'+'DNN'+'-'+sampling+'-'+'Validation-Table.xlsx')
